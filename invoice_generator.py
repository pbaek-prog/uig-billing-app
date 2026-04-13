"""Generate professional Excel invoices for US Immigration Group legal billing."""
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "generated_invoices")
os.makedirs(OUTPUT_DIR, exist_ok=True)

FIRM = {
    "name": "US IMMIGRATION GROUP",
    "address": "800 E. Northwest Hwy. Ste 205",
    "city_state": "Mount Prospect, IL 60016",
    "phone": "(847) 449-8660",
    "email": "info@usimmigrationgroup.org",
    "website": "www.usimmigrationgroup.org",
}

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
GOLD = "F5A623"
NAVY = "1A3C5E"
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
TOTAL_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
SUBTOTAL_FILL = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")


def generate_invoice_excel(client, invoice_number, invoice_date, due_date,
                           description, legal_fees=0, filing_fees=0,
                           other_expenses=0, retainer_applied=0, line_items=None):
    """Generate a legal invoice as an Excel file.

    Args:
        client: dict with keys name, email, address, case_number, visa_type
        invoice_number: str e.g. "INV-2026-001"
        invoice_date: date object
        due_date: date object
        description: str case description
        legal_fees: float
        filing_fees: float
        other_expenses: float
        retainer_applied: float
        line_items: list of (description, amount) tuples (overrides auto-generation)

    Returns:
        Path to generated file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    title_font = Font(name="Arial", bold=True, size=22, color=NAVY)
    firm_font = Font(name="Arial", bold=True, size=16, color=NAVY)
    label_font = Font(name="Arial", bold=True, size=11)
    normal_font = Font(name="Arial", size=11)
    small_font = Font(name="Arial", size=10, color="666666")
    gold_line = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")

    # Gold line at top
    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = gold_line
    ws.row_dimensions[1].height = 4

    # Firm header
    ws.merge_cells('B2:C2')
    ws['B2'] = FIRM["name"]
    ws['B2'].font = firm_font
    ws['B3'] = FIRM["address"]
    ws['B3'].font = normal_font
    ws['B4'] = FIRM["city_state"]
    ws['B4'].font = normal_font
    ws['B5'] = f"Tel: {FIRM['phone']}  |  {FIRM['email']}"
    ws['B5'].font = small_font
    ws['B6'] = FIRM["website"]
    ws['B6'].font = small_font

    # INVOICE title
    ws['D2'] = "INVOICE"
    ws['D2'].font = title_font
    ws['D2'].alignment = Alignment(horizontal='right')

    # Gold line below header
    for col in range(1, 5):
        ws.cell(row=7, column=col).fill = gold_line
    ws.row_dimensions[7].height = 4

    # Invoice details
    details = [
        (9, "Invoice #:", invoice_number),
        (10, "Date:", invoice_date.strftime("%B %d, %Y") if isinstance(invoice_date, date) else str(invoice_date)),
        (11, "Due Date:", due_date.strftime("%B %d, %Y") if isinstance(due_date, date) else str(due_date)),
        (12, "Matter:", description or "Legal Services"),
    ]
    for row, lbl, val in details:
        ws.cell(row=row, column=3, value=lbl).font = label_font
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
        ws.cell(row=row, column=4, value=val).font = normal_font

    # Bill To
    ws['B9'] = "BILL TO:"
    ws['B9'].font = Font(name="Arial", bold=True, size=11, color=NAVY)
    ws['B10'] = client.get("name", "")
    ws['B10'].font = Font(name="Arial", bold=True, size=12)
    if client.get("address"):
        ws['B11'] = client["address"]
        ws['B11'].font = normal_font
    case_info = []
    if client.get("case_number"):
        case_info.append(f"Case: {client['case_number']}")
    if client.get("visa_type"):
        case_info.append(f"Type: {client['visa_type']}")
    if case_info:
        ws['B12'] = "  |  ".join(case_info)
        ws['B12'].font = small_font

    # Line items header
    row = 14
    for col, header in [(2, 'DESCRIPTION'), (3, 'CATEGORY'), (4, 'AMOUNT')]:
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center' if col > 2 else 'left')
        cell.border = THIN_BORDER

    # Build line items
    if line_items is None:
        line_items = []
        if legal_fees > 0:
            line_items.append((f"Legal Services — {description or 'Legal Fees'}", "Legal Fees", legal_fees))
        if filing_fees > 0:
            line_items.append(("USCIS Filing Fees", "Filing Fees", filing_fees))
        if other_expenses > 0:
            line_items.append(("Additional Expenses (translations, postage, etc.)", "Expenses", other_expenses))

    row = 15
    for item in line_items:
        desc, cat, amount = item[0], item[1], item[2]
        ws.cell(row=row, column=2, value=desc).font = normal_font
        ws.cell(row=row, column=3, value=cat).font = normal_font
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        a_cell = ws.cell(row=row, column=4, value=amount)
        a_cell.font = normal_font
        a_cell.number_format = '$#,##0.00'
        a_cell.alignment = Alignment(horizontal='right')
        for c in range(2, 5):
            ws.cell(row=row, column=c).border = THIN_BORDER
        row += 1

    # Pad empty rows
    for r in range(row, row + max(0, 6 - len(line_items))):
        for c in range(2, 5):
            ws.cell(row=r, column=c).border = THIN_BORDER
        row = r + 1

    # Totals section
    total = legal_fees + filing_fees + other_expenses
    amount_due = total - retainer_applied

    row += 1
    totals = [
        ("Total Legal Fees", legal_fees),
        ("Total Filing Fees & Expenses", filing_fees + other_expenses),
        ("Subtotal", total),
    ]
    if retainer_applied > 0:
        totals.append(("Less: Retainer Applied", -retainer_applied))
    totals.append(("AMOUNT DUE", amount_due))

    for label, value in totals:
        ws.cell(row=row, column=3, value=label).font = label_font
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='right')
        val_cell = ws.cell(row=row, column=4, value=abs(value) if value < 0 else value)
        val_cell.number_format = '$#,##0.00' if value >= 0 else '($#,##0.00)'
        val_cell.alignment = Alignment(horizontal='right')

        if label == "AMOUNT DUE":
            for c in [3, 4]:
                ws.cell(row=row, column=c).fill = TOTAL_FILL
                ws.cell(row=row, column=c).font = TOTAL_FONT
        elif label == "Subtotal":
            for c in [3, 4]:
                ws.cell(row=row, column=c).fill = SUBTOTAL_FILL
            val_cell.font = label_font
        elif "Retainer" in label:
            val_cell.font = Font(name="Arial", size=11, color="CC0000")
            val_cell.number_format = '($#,##0.00)'
        else:
            val_cell.font = normal_font
        row += 1

    # Payment instructions
    row += 2
    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'] = "Payment Information"
    ws[f'B{row}'].font = Font(name="Arial", bold=True, size=11, color=NAVY)
    row += 1
    instructions = [
        f"Payment Due: Upon Receipt",
        f"Check: Payable to US Immigration Group",
        f"Zelle: {FIRM['email']}",
        f"Wire Transfer: Contact office for details",
    ]
    for line in instructions:
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'] = line
        ws[f'B{row}'].font = small_font
        row += 1

    # Footer
    row += 1
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = gold_line
    ws.row_dimensions[row].height = 4
    row += 1
    ws.merge_cells(f'B{row}:D{row}')
    ws[f'B{row}'] = "Thank you for choosing US Immigration Group."
    ws[f'B{row}'].font = Font(name="Arial", italic=True, size=10, color="666666")
    ws[f'B{row}'].alignment = Alignment(horizontal='center')

    # Save
    safe_name = client.get("name", "Client").replace(",", "").replace(".", "").replace("/", "_").replace(" ", "_")[:25]
    filename = f"Invoice_{invoice_number}_{safe_name}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    return filepath
